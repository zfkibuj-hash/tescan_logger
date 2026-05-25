"""Exporters for CSV, Excel (openpyxl), and PDF (reportlab).

Exports session data, vacuum cycles, and billing summaries.
"""

import csv
import logging
import os
from datetime import datetime
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)


class CSVExporter:
    """Export data to CSV files."""

    def export_sessions(self, sessions: List[Dict], output_path: str) -> str:
        """Export session data to CSV."""
        if not sessions:
            logger.warning("No sessions to export")
            return ""

        fieldnames = [
            "id", "username", "start_time", "end_time",
            "duration_seconds", "gvl_total_seconds", "gvl_cycle_count",
            "status", "cost", "discount_percent",
            "override_cost", "override_time_minutes",
            "excluded_from_billing", "source_file", "notes",
        ]

        try:
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f, fieldnames=fieldnames, extrasaction="ignore"
                )
                writer.writeheader()
                for session in sessions:
                    writer.writerow(session)
            logger.info("Exported %d sessions to %s", len(sessions), output_path)
            return output_path
        except OSError as e:
            logger.error("CSV export failed: %s", e)
            return ""

    def export_vacuum_cycles(
        self, cycles: List[Dict], output_path: str
    ) -> str:
        """Export vacuum cycle data to CSV."""
        if not cycles:
            return ""

        fieldnames = [
            "id", "session_id", "pump_start", "ready_time",
            "end_time", "status", "pump_duration_seconds", "source_file",
        ]

        try:
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(
                    f, fieldnames=fieldnames, extrasaction="ignore"
                )
                writer.writeheader()
                for cycle in cycles:
                    writer.writerow(cycle)
            logger.info("Exported %d vacuum cycles to %s", len(cycles), output_path)
            return output_path
        except OSError as e:
            logger.error("CSV export failed: %s", e)
            return ""

    def export_billing_summary(
        self, summary: Dict[str, Any], output_path: str
    ) -> str:
        """Export billing summary to CSV."""
        try:
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Metric", "Value"])
                for key, value in summary.items():
                    writer.writerow([key, value])
            return output_path
        except OSError as e:
            logger.error("CSV export failed: %s", e)
            return ""


class ExcelExporter:
    """Export data to Excel using openpyxl."""

    def export_sessions(
        self,
        sessions: List[Dict],
        output_path: str,
        title: str = "Sessions Report",
    ) -> str:
        """Export session data to Excel workbook."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl not installed - cannot export to Excel")
            return ""

        if not sessions:
            logger.warning("No sessions to export")
            return ""

        wb = Workbook()
        ws = wb.active
        ws.title = "Sessions"

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
        )

        headers = [
            "ID", "Username", "Start Time", "End Time",
            "Duration (s)", "GVL Total (s)", "GVL Cycles",
            "Status", "Cost (PLN)", "Discount %",
            "Notes",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, session in enumerate(sessions, start=2):
            ws.cell(row=row_idx, column=1, value=session.get("id"))
            ws.cell(row=row_idx, column=2, value=session.get("username"))
            ws.cell(row=row_idx, column=3, value=session.get("start_time"))
            ws.cell(row=row_idx, column=4, value=session.get("end_time"))
            ws.cell(row=row_idx, column=5, value=session.get("duration_seconds"))
            ws.cell(row=row_idx, column=6, value=session.get("gvl_total_seconds"))
            ws.cell(row=row_idx, column=7, value=session.get("gvl_cycle_count"))
            ws.cell(row=row_idx, column=8, value=session.get("status"))
            ws.cell(row=row_idx, column=9, value=session.get("cost"))
            ws.cell(row=row_idx, column=10, value=session.get("discount_percent"))
            ws.cell(row=row_idx, column=11, value=session.get("notes", ""))

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        try:
            wb.save(output_path)
            logger.info("Exported %d sessions to %s", len(sessions), output_path)
            return output_path
        except OSError as e:
            logger.error("Excel export failed: %s", e)
            return ""

    def export_full_report(
        self,
        sessions: List[Dict],
        vacuum_cycles: List[Dict],
        penalties: List[Dict],
        summary: Dict[str, Any],
        output_path: str,
    ) -> str:
        """Export full report with multiple sheets."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed")
            return ""

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.cell(row=1, column=1, value="TESCAN VEGA3 - Billing Report")
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary.cell(
            row=2, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        )

        row = 4
        for key, value in summary.items():
            ws_summary.cell(row=row, column=1, value=key)
            ws_summary.cell(row=row, column=2, value=str(value))
            row += 1

        # Sessions sheet
        ws_sessions = wb.create_sheet("Sessions")
        if sessions:
            headers = list(sessions[0].keys())
            for col, h in enumerate(headers, 1):
                ws_sessions.cell(row=1, column=col, value=h)
            for r_idx, session in enumerate(sessions, 2):
                for c_idx, key in enumerate(headers, 1):
                    ws_sessions.cell(row=r_idx, column=c_idx, value=session.get(key))

        # Vacuum sheet
        ws_vacuum = wb.create_sheet("Vacuum Cycles")
        if vacuum_cycles:
            headers = list(vacuum_cycles[0].keys())
            for col, h in enumerate(headers, 1):
                ws_vacuum.cell(row=1, column=col, value=h)
            for r_idx, cycle in enumerate(vacuum_cycles, 2):
                for c_idx, key in enumerate(headers, 1):
                    ws_vacuum.cell(row=r_idx, column=c_idx, value=cycle.get(key))

        try:
            wb.save(output_path)
            logger.info("Full report exported to %s", output_path)
            return output_path
        except OSError as e:
            logger.error("Excel export failed: %s", e)
            return ""


class PDFExporter:
    """Export data to PDF using reportlab."""

    def export_sessions(
        self,
        sessions: List[Dict],
        output_path: str,
        title: str = "TESCAN VEGA3 - Sessions Report",
    ) -> str:
        """Export session data to PDF."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
            )
        except ImportError:
            logger.error("reportlab not installed - cannot export to PDF")
            return ""

        if not sessions:
            logger.warning("No sessions to export")
            return ""

        try:
            doc = SimpleDocTemplate(
                output_path,
                pagesize=landscape(A4),
                leftMargin=30,
                rightMargin=30,
            )
            styles = getSampleStyleSheet()
            elements = []

            # Title
            elements.append(Paragraph(title, styles["Title"]))
            elements.append(Spacer(1, 12))
            elements.append(
                Paragraph(
                    f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    styles["Normal"],
                )
            )
            elements.append(Spacer(1, 20))

            # Table data
            headers = [
                "ID", "User", "Start", "End",
                "GVL (s)", "Cycles", "Status", "Cost (PLN)",
            ]
            table_data = [headers]

            for s in sessions:
                row = [
                    str(s.get("id", "")),
                    s.get("username", ""),
                    str(s.get("start_time", ""))[:16],
                    str(s.get("end_time", ""))[:16],
                    f"{s.get('gvl_total_seconds', 0):.0f}",
                    str(s.get("gvl_cycle_count", 0)),
                    s.get("status", ""),
                    f"{s.get('cost', 0):.2f}",
                ]
                table_data.append(row)

            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3366CC")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0F0")]),
            ]))
            elements.append(table)

            doc.build(elements)
            logger.info("Exported %d sessions to PDF: %s", len(sessions), output_path)
            return output_path
        except Exception as e:
            logger.error("PDF export failed: %s", e)
            return ""

    def export_billing_summary(
        self,
        summary: Dict[str, Any],
        sessions: List[Dict],
        output_path: str,
    ) -> str:
        """Export billing summary to PDF."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
            )
        except ImportError:
            logger.error("reportlab not installed")
            return ""

        try:
            doc = SimpleDocTemplate(output_path, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(
                Paragraph("TESCAN VEGA3 - Billing Summary", styles["Title"])
            )
            elements.append(Spacer(1, 20))

            # Summary table
            summary_data = [["Metric", "Value"]]
            for key, value in summary.items():
                label = key.replace("_", " ").title()
                summary_data.append([label, str(value)])

            summary_table = Table(summary_data, colWidths=[250, 150])
            summary_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3366CC")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]))
            elements.append(summary_table)

            doc.build(elements)
            logger.info("Billing summary exported to PDF: %s", output_path)
            return output_path
        except Exception as e:
            logger.error("PDF export failed: %s", e)
            return ""
